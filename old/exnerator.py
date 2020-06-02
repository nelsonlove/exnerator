#!/usr/bin/python
#Exnerator 2.1

#nonlocal variables
version = 2.1
global detailedParse

#import our stuff
import sys
import openpyxl as xl
#import test
import record
import art
from os import system, name, path
import random
from texttable import Texttable
import codes
import structuralSummary
import time

#nonlocal variables
  
# define our clear function 
def clear(): 
    # for windows 
    if name == 'nt': 
        _ = system('cls') 
    # for mac and linux(here, os.name is 'posix') 
    else: 
        _ = system('clear') 

def splash():
	clear()
	r = random.randint(0, 1)
	fonts = ["speed","funfaces"]
	art.tprint("EXNERATOR",font=fonts[r])
	print("---------------")
	print("Exnerator v",version)
	print("---------------\n")

def yes_or_no(question):
    while "Invalid response!":
        reply = str(input(question+' (y/n): ')).lower().strip()
        if reply == 'y':
            return True
        if reply == 'n':
            return False
        else:
        	pass

def menu(prompt,responses):
	print(prompt)
	r = 1
	v = []
	for response in responses:
		print(r,". ",response,sep='')
		v.append(r)
		r += 1
	while "Invalid response!":
		reply = str(input('Enter option ['+str(min(v))+'-'+str(max(v))+', q to quit] : ')).strip()
		if reply.isdigit() and int(reply) in v:
			return int(reply)
		elif reply == "q":
			return 0

def loadFromExcel(filename):
	excelData = []
	wb = xl.load_workbook(filename)
	print('File "', filename, '" loaded.',sep='')
	ws = wb.active
	print('Sheet "', ws.title, '" loaded.',sep='')
	numResponses = ws.max_row - 1
	print(numResponses, "responses found.")
	for row in range(numResponses):
		row += 1
		data = []
		for cell in ws[row+1]:
			data.append(cell.value)
		excelData.append(data)
		if row == 1 and len(excelData[0]) != 9:
			print("Invalid number of columns!")
			quit()
	print("Loaded",len(excelData),"responses.")
	return excelData

def chooseFile(message):
	if message != None: print(message)
	while True:
		filename = input("Enter a filename: ")
		if path.isfile(filename):
			if filename.endswith(".xlsx"):
				return filename
			else:
				print("File must be .xlsx format!")
		else:
			print("File not found!")

def main():
	splash()
	if len(sys.argv) < 2:
		filename = chooseFile("No file specified!")
	elif len(sys.argv) > 2:
		filename = chooseFile("Too many arguments!")
	else:
		filename = sys.argv[1]

	#	r = test.Record(filename)

	data = loadFromExcel(filename)
	r = record.Record(data)

	if False:#yes_or_no("Print response table?"):
		r.printTable()
	if r.errors == 0:
		print("Successfully parsed record!")
		if True: #if yes_or_no("Calculate structural summary?"):
			s = structuralSummary.calc(r)
	else:
		print(r.errors,"errors found. Please review data and try again.")

if __name__ == "__main__":
	main()