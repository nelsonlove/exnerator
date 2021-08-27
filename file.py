from pathlib import Path
import openpyxl as xl


class InvalidColumnsError(Exception):
    def __init__(self, file, sheet):
        self.file = file
        self.sheet = sheet

        super().__init__()

    def __str__(self):
        return f'Sheet {self.sheet.max_column} in file {self.file} has {self.sheet.max_column} columns; expected 9.'


class File:
    @classmethod
    def load_from_excel(cls, filename):
        book = xl.load_workbook(filename)
        print('File "', filename, '" loaded.', sep='')
        sheet = book.active
        if sheet.max_column != 9:
            raise InvalidColumnsError(filename, sheet)
        print(f'Sheet {sheet.title} loaded.')
        # print([cell.value for cell in sheet[1]])
        # quit()
        num_responses = sheet.max_row - 1
        print(f'{num_responses} responses found.')
        out = []
        for row in range(2, num_responses + 2):
            out.append([cell.value for cell in sheet[row]])
        return out


# Takes an Excel file and returns a 2d array.
#
# import openpyxl as xl
#
# import parse_old as Parse
#
#
# def Excel(filename):
#     def load(filename):
#         book = xl.load_workbook(filename)
#         print('File "', filename, '" loaded.', sep='')
#         sheet = book.active
#         if sheet.max_column != 9:
#             print("Invalid number of columns!")
#             quit()
#         print('Sheet "', sheet.title, '" loaded.', sep='')
#         n = sheet.max_row - 1
#         print(n, "responses found.")
#         out = []
#         for row in range(2, n + 2):
#             data = []
#             for cell in sheet[row]:
#                 data.append(cell.value)
#             out.append(data)
#         return out
#
#     def parse(data, n):
#         score = {}
#         columns = [
#             Parse.card(row[0]),
#             Parse.locDev(row[1]),
#             Parse.locNum(row[2]),
#             Parse.detFQ(row[3]),
#             Parse.pair(row[4]),
#             Parse.contents(row[5]),
#             Parse.popular(row[6]),
#             Parse.zScore(row[7]),
#             Parse.special(row[8])]
#         for index, column in enumerate(columns):
#             try:
#                 score.update(column)
#             except:
#                 y = str(n)
#                 x = index + 1
#                 print("Invalid data at cell " + chr(ord('@') + x) + y + ': "' + str(column) + '"')
#         return score
#
#     rows = load(filename)
#     out = []
#     for index, row in enumerate(rows):
#         out.append(parse(row, index))
#     return out
