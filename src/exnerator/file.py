import openpyxl as xl

from .exceptions import InvalidColumnsError, NoResponsesError
from .models import File


class ExcelFile(File):
    def __init__(self, filename):
        super().__init__(filename)

    def __str__(self):
        return f'Excel file {self.filename}'

    def open(self):
        book = xl.load_workbook(self.filename)
        sheet = book.active
        print(f'Sheet {sheet.title} loaded.')
        return sheet

    def validate(self):
        if self.columns != 9:
            raise InvalidColumnsError(self)
        elif self.responses < 1:
            raise NoResponsesError(self)

    @property
    def columns(self):
        return self._data.max_column

    @property
    def responses(self):
        return self._data.max_row - 1

    @property
    def data(self):
        out = []
        for row in range(2, self.responses + 2):
            out.append([cell.value for cell in self._data[row]])
        return out


# Takes an Excel file and returns a 2d array.
#
# import openpyxl as xl
#
# import parse_old as Parse
#
#
# def Excel(filename):
#     def load(filename):
#         book = xl.load_workbook(filename)
#         print('File "', filename, '" loaded.', sep='')
#         sheet = book.active
#         if sheet.max_column != 9:
#             print("Invalid number of columns!")
#             quit()
#         print('Sheet "', sheet.title, '" loaded.', sep='')
#         n = sheet.max_row - 1
#         print(n, "responses found.")
#         out = []
#         for row in range(2, n + 2):
#             data = []
#             for cell in sheet[row]:
#                 data.append(cell.value)
#             out.append(data)
#         return out
#
#     def parse(data, n):
#         score = {}
#         columns = [
#             Parse.card(row[0]),
#             Parse.locDev(row[1]),
#             Parse.locNum(row[2]),
#             Parse.detFQ(row[3]),
#             Parse.pair(row[4]),
#             Parse.contents(row[5]),
#             Parse.popular(row[6]),
#             Parse.zScore(row[7]),
#             Parse.special(row[8])]
#         for index, column in enumerate(columns):
#             try:
#                 score.update(column)
#             except:
#                 y = str(n)
#                 x = index + 1
#                 print("Invalid data at cell " + chr(ord('@') + x) + y + ': "' + str(column) + '"')
#         return score
#
#     rows = load(filename)
#     out = []
#     for index, row in enumerate(rows):
#         out.append(parse(row, index))
#     return out
